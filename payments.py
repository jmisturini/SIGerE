import os
import math
from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app
from flask_login import login_required, current_user
from models import User, Course, TeacherBasePay, TeacherAdditivePayment, TeacherOvertimePay
from forms import FormTeacherBasePay, FormTeacherAdditivePay, FormTeacherOvertimePay
from extensions import db
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from io import BytesIO
from decimal import Decimal

bp = Blueprint('payments', __name__, url_prefix='/payments')

# ================= HELPER FUNCTIONS (Signals Logic) =================

def get_remaining_months(month_base):
    """Calculate remaining months in the semester."""
    month = int(month_base.split('-')[1])
    first_semester = [2, 3, 4, 5, 6, 7]
    second_semester = [8, 9, 10, 11, 12, 1]
    
    if month in first_semester:
        return len([m for m in first_semester if m >= month])
    else:
        if month == 1: return 1
        return len([m for m in second_semester if m >= month])

def calculate_base_hours(pay_obj):
    """Pre-save logic for Base and Additive payments."""
    if isinstance(pay_obj, TeacherBasePay):
        base_hour = int(pay_obj.weekly_workload)
    else:
        base_hour = int(pay_obj.additional_hour)
        
    remaining_months = get_remaining_months(pay_obj.month_start)
    month = base_hour * 4.5 if base_hour else 0
    semester = month * remaining_months if base_hour else 0
    
    pay_obj.monthly_hour = math.ceil(month)
    pay_obj.semester_hour = math.ceil(semester)
    pay_obj.accountable_id = current_user.id

# ================= VALIDATIONS (Decorators Logic) =================

def validate_base_pay_create(form, teacher_id, month_start, month_end):
    # Check duplicates
    exists = TeacherBasePay.query.filter(
        TeacherBasePay.month_start >= month_start,
        TeacherBasePay.month_end <= month_end,
        TeacherBasePay.teacher_id == teacher_id
    ).first()
    if exists:
        flash('Erro: Lançamento duplicado para esse semestre.', 'danger')
        return False
    
    # Check past month
    month_val = int(month_start.split('-')[1])
    if month_val < datetime.now().month:
        flash('Erro: Esse mês não pode ser lançado!', 'danger')
        return False
        
    return True

def validate_30_days_rule(created_at):
    if created_at.month < datetime.now().month or (datetime.now().date() - created_at.date() > timedelta(days=30)):
        flash('Erro: Registros com mais de 30 dias não podem ser alterados!', 'danger')
        return False
    return True

def validate_180_days_rule(created_at):
    if datetime.now().date() - created_at.date() > timedelta(days=180):
        flash('Erro: Esse registro não pode ser excluído/alterado (mais de 180 dias).', 'danger')
        return False
    return True

# ================= BASE PAY ROUTES =================

@bp.route('/list')
@login_required
def list_base_pays():
    infos = TeacherBasePay.query.order_by(TeacherBasePay.created_at.desc()).all()
    return render_template('payments/list_base.html', infos=infos)

@bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_base_pay():
    form = FormTeacherBasePay()
    form.teacher.choices = [(t.id, t.full_name) for t in User.query.filter_by(profile_type='teacher', is_active_user=True).order_by(User.full_name).all()]
    form.course.choices = [(c.id, c.name) for c in Course.query.filter_by(is_active=True).order_by(Course.name).all()]
    
    if form.validate_on_submit():
        if not validate_base_pay_create(form, form.teacher.data, form.month_start.data, form.month_end.data):
            return redirect(url_for('payments.create_base_pay'))
            
        pay = TeacherBasePay(
            teacher_id=form.teacher.data, course_id=form.course.data,
            month_start=form.month_start.data, month_end=form.month_end.data,
            budget_code=form.budget_code.data, complement=form.complement.data,
            weekly_workload=form.weekly_workload.data
        )
        calculate_base_hours(pay)
        db.session.add(pay)
        db.session.commit()
        flash('Lançamento Base realizado!', 'success')
        return redirect(url_for('payments.list_base_pays'))
        
    return render_template('payments/form_base.html', form=form, title='Novo Lançamento Base')

@bp.route('/edit/<int:pay_id>', methods=['GET', 'POST'])
@login_required
def edit_base_pay(pay_id):
    pay = TeacherBasePay.query.get_or_404(pay_id)
    if not validate_30_days_rule(pay.created_at):
        return redirect(url_for('payments.list_base_pays'))
        
    form = FormTeacherBasePay(obj=pay)
    form.teacher.choices = [(t.id, t.full_name) for t in User.query.filter_by(profile_type='teacher', is_active_user=True).order_by(User.full_name).all()]
    form.course.choices = [(c.id, c.name) for c in Course.query.filter_by(is_active=True).order_by(Course.name).all()]
    
    if form.validate_on_submit():
        pay.teacher_id = form.teacher.data
        pay.course_id = form.course.data
        pay.month_start = form.month_start.data
        pay.month_end = form.month_end.data
        pay.budget_code = form.budget_code.data
        pay.complement = form.complement.data
        pay.weekly_workload = form.weekly_workload.data
        calculate_base_hours(pay)
        db.session.commit()
        flash('Alteração realizada!', 'success')
        return redirect(url_for('payments.list_base_pays'))
        
    return render_template('payments/form_base.html', form=form, title='Editar Lançamento Base')

@bp.route('/delete/<int:pay_id>', methods=['POST'])
@login_required
def delete_base_pay(pay_id):
    pay = TeacherBasePay.query.get_or_404(pay_id)
    if not validate_180_days_rule(pay.created_at):
        return redirect(url_for('payments.list_base_pays'))
        
    db.session.delete(pay)
    db.session.commit()
    flash('Registro excluído com sucesso', 'success')
    return redirect(url_for('payments.list_base_pays'))

# ================= ADDITIVE ROUTES =================

@bp.route('/additive/create', methods=['GET', 'POST'])
@login_required
def create_additive():
    form = FormTeacherAdditivePay()
    bases = TeacherBasePay.query.order_by(TeacherBasePay.created_at.desc()).all()
    form.base_release.choices = [(b.id, f"{b.teacher.full_name} - {b.month_start}") for b in bases]
    form.course.choices = [(c.id, c.name) for c in Course.query.filter_by(is_active=True).order_by(Course.name).all()]
    
    if form.validate_on_submit():
        base = TeacherBasePay.query.get(form.base_release.data)
        if not validate_180_days_rule(base.created_at):
            return redirect(url_for('payments.list_base_pays'))
            
        # Check duplicate month
        exists = TeacherAdditivePayment.query.filter_by(base_release_id=base.id, month_start=form.month_start.data).first()
        if exists:
            flash('Erro: Professor já tem aditivo para esse mês.', 'danger')
            return redirect(url_for('payments.create_additive'))
            
        additive = TeacherAdditivePayment(
            base_release_id=form.base_release.data, course_id=form.course.data,
            month_start=form.month_start.data, month_end=form.month_end.data,
            additional_hour=form.additional_hour.data, complement=form.complement.data
        )
        calculate_base_hours(additive)
        db.session.add(additive)
        db.session.commit()
        flash('Aditivo realizado!', 'success')
        return redirect(url_for('payments.list_base_pays'))
        
    return render_template('payments/form_additive.html', form=form, title='Novo Aditivo')

@bp.route('/additive/delete/<int:additive_id>', methods=['POST'])
@login_required
def delete_additive(additive_id):
    additive = TeacherAdditivePayment.query.get_or_404(additive_id)
    if not validate_180_days_rule(additive.created_at):
        return redirect(url_for('payments.list_base_pays'))
        
    db.session.delete(additive)
    db.session.commit()
    flash('Aditivo excluído com sucesso', 'success')
    return redirect(url_for('payments.list_base_pays'))

# ================= OVERTIME ROUTES =================

@bp.route('/overtime/list')
@login_required
def list_overtime():
    filter_month = request.args.get('month_base', '')
    filter_teacher = request.args.get('teacher_filter', type=int)
    
    query = TeacherOvertimePay.query
    
    # Aplica os filtros se forem fornecidos
    if filter_month:
        query = query.filter_by(month_base=filter_month)
    if filter_teacher:
        query = query.filter_by(teacher_id=filter_teacher)
        
    infos = query.order_by(TeacherOvertimePay.created_at.desc()).all()
    
    # Busca professores ativos para popular o dropdown
    list_teachers = User.query.filter_by(profile_type='teacher', is_active_user=True).order_by(User.full_name).all()
    
    return render_template('payments/list_overtime.html', 
                           infos=infos, 
                           list_teachers=list_teachers,
                           filter_month=filter_month,
                           filter_teacher=filter_teacher)

@bp.route('/overtime/create', methods=['GET', 'POST'])
@login_required
def create_overtime():
    form = FormTeacherOvertimePay()
    form.teacher.choices = [(t.id, t.full_name) for t in User.query.filter_by(profile_type='teacher', is_active_user=True).order_by(User.full_name).all()]
    
    # Pré-preenche o mês base com o mês e ano atuais
    if request.method == 'GET':
        form.month_base.data = datetime.now().strftime('%Y-%m')
    
    if form.validate_on_submit():
        import re
        current_date = datetime.now()
        month_base_str = form.month_base.data
        
        if not re.match(r'^\d{4}-\d{2}$', month_base_str):
            flash('Erro: O formato do Mês Base é inválido. Utilize YYYY-MM (ex: 2024-05).', 'danger')
            return redirect(url_for('payments.create_overtime'))
            
        try:
            month_base = datetime.strptime(month_base_str, '%Y-%m')
        except ValueError:
            flash('Erro: O Mês Base inserido não é uma data válida.', 'danger')
            return redirect(url_for('payments.create_overtime'))

        if month_base.year == current_date.year and month_base.month == current_date.month and current_date.day > 25:
            flash('Erro: Lançamentos do mês atual só podem ser feitos até o dia 25.', 'danger')
            return redirect(url_for('payments.create_overtime'))
            
        try:
            hourly_value = float(form.hourly_value.data.replace('.', '').replace(',', '.'))
            # Validação: Valor H/a deve ser maior que 0
            if hourly_value <= 0:
                flash('Erro: O Valor H/a deve ser maior que 0.', 'danger')
                return redirect(url_for('payments.create_overtime'))
        except ValueError:
            flash('Erro: Insira um valor válido para "Valor H/a"', 'danger')
            return redirect(url_for('payments.create_overtime'))
            
        overtime = TeacherOvertimePay(
            teacher_id=form.teacher.data, teaching_level=form.teaching_level.data,
            weekly_workload=form.weekly_workload.data, hourly_value=hourly_value,
            budget_code=form.budget_code.data, shift=form.shift.data,
            multiple_dates=form.multiple_dates.data, justification=form.justification.data,
            month_base=form.month_base.data, accountable_id=current_user.id
        )
        db.session.add(overtime)
        db.session.commit()
        flash('Lançamento de Hora Extra realizado!', 'success')
        return redirect(url_for('payments.list_overtime'))
        
    return render_template('payments/form_overtime.html', form=form, title='Nova Hora Extra')

@bp.route('/overtime/edit/<int:overtime_id>', methods=['GET', 'POST'])
@login_required
def edit_overtime(overtime_id):
    overtime = TeacherOvertimePay.query.get_or_404(overtime_id)
    
    if overtime.created_at.month < datetime.now().month or (datetime.now().date() - overtime.created_at.date() > timedelta(days=30)):
        flash('Erro: Registros dos meses anteriores não podem ser alterados.', 'danger')
        return redirect(url_for('payments.list_overtime'))

    form = FormTeacherOvertimePay(obj=overtime)
    form.teacher.choices = [(t.id, t.full_name) for t in User.query.filter_by(profile_type='teacher', is_active_user=True).order_by(User.full_name).all()]
    
    if request.method == 'GET':
        form.hourly_value.data = f"{overtime.hourly_value:.2f}".replace('.', ',')

    if form.validate_on_submit():
        import re
        month_base_str = form.month_base.data
        
        if not re.match(r'^\d{4}-\d{2}$', month_base_str):
            flash('Erro: O formato do Mês Base é inválido. Utilize YYYY-MM (ex: 2024-05).', 'danger')
            return redirect(url_for('payments.edit_overtime', overtime_id=overtime_id))
            
        try:
            datetime.strptime(month_base_str, '%Y-%m')
        except ValueError:
            flash('Erro: O Mês Base inserido não é uma data válida.', 'danger')
            return redirect(url_for('payments.edit_overtime', overtime_id=overtime_id))

        try:
            hourly_value = float(form.hourly_value.data.replace('.', '').replace(',', '.'))
            # Validação: Valor H/a deve ser maior que 0
            if hourly_value <= 0:
                flash('Erro: O Valor H/a deve ser maior que 0.', 'danger')
                return redirect(url_for('payments.edit_overtime', overtime_id=overtime_id))
        except ValueError:
            flash('Erro: Insira um valor válido para "Valor H/a"', 'danger')
            return redirect(url_for('payments.edit_overtime', overtime_id=overtime_id))

        overtime.teacher_id = form.teacher.data
        overtime.teaching_level = form.teaching_level.data
        overtime.weekly_workload = form.weekly_workload.data
        overtime.hourly_value = hourly_value
        overtime.budget_code = form.budget_code.data
        overtime.shift = form.shift.data
        overtime.multiple_dates = form.multiple_dates.data
        overtime.justification = form.justification.data
        overtime.month_base = form.month_base.data
        overtime.accountable_id = current_user.id

        db.session.commit()
        flash('Alteração realizada!', 'success')
        return redirect(url_for('payments.list_overtime'))

        overtime.teacher_id = form.teacher.data
        overtime.teaching_level = form.teaching_level.data
        overtime.weekly_workload = form.weekly_workload.data
        overtime.hourly_value = hourly_value
        overtime.budget_code = form.budget_code.data
        overtime.shift = form.shift.data
        overtime.multiple_dates = form.multiple_dates.data
        overtime.justification = form.justification.data
        overtime.month_base = form.month_base.data
        overtime.accountable_id = current_user.id

        db.session.commit()
        flash('Alteração realizada!', 'success')
        return redirect(url_for('payments.list_overtime'))

    return render_template('payments/form_overtime.html', form=form, title='Editar Hora Extra')

@bp.route('/overtime/delete/<int:overtime_id>', methods=['POST'])
@login_required
def delete_overtime(overtime_id):
    overtime = TeacherOvertimePay.query.get_or_404(overtime_id)
    # Validate 30 days or past month
    if overtime.created_at.month < datetime.now().month or (datetime.now().date() - overtime.created_at.date() > timedelta(days=30)):
        flash('Erro: Registros dos meses anteriores não podem ser excluídos.', 'danger')
        return redirect(url_for('payments.list_overtime'))
        
    db.session.delete(overtime)
    db.session.commit()
    flash('Registro de Hora Extra excluído', 'success')
    return redirect(url_for('payments.list_overtime'))

# ================= EXCEL EXPORT ROUTES (USANDO MODELOS ORIGINAIS) =================

@bp.route('/export/base')
@login_required
def export_excel_base():
    semester_base = request.args.get('semester_base', type=int)
    year_base = request.args.get('year_base')
    
    if not semester_base or not year_base:
        flash('Selecione o semestre e ano.', 'danger')
        return redirect(url_for('payments.list_base_pays'))
        
    if semester_base == 1:
        start_q, end_q = f'{year_base}-02', f'{year_base}-07'
    else:
        start_q, end_q = f'{year_base}-08', f'{int(year_base)+1}-01'

    pays = TeacherBasePay.query.filter(
        TeacherBasePay.month_start >= start_q,
        TeacherBasePay.month_end <= end_q
    ).all()

    if not pays:
        flash('Informações não encontradas.', 'danger')
        return redirect(url_for('payments.list_base_pays'))

    # Caminho para o arquivo modelo
    template_path = os.path.join(current_app.root_path, 'static', 'templates_excel', 'base_planilha_CH_semestral_professores.xlsx')
    
    if not os.path.exists(template_path):
        flash('Modelo do Excel não encontrado no servidor.', 'danger')
        return redirect(url_for('payments.list_base_pays'))

    workbook = load_workbook(filename=template_path)
    ws = workbook['CH Mensal']
    
    list_table = []

    # Prepara os dados e verifica aditivos
    for pay in pays:
        dic_info = {
            'id_teacher': pay.teacher.registration or 'N/A',
            'name_teacher': pay.teacher.full_name,
            'monthly_hour_query': pay.monthly_hour,
            'additive': False
        }
        
        # Soma as horas do aditivo se existir no semestre
        additives = TeacherAdditivePayment.query.filter_by(base_release_id=pay.id).all()
        for additive in additives:
            dic_info['monthly_hour_query'] += additive.monthly_hour
            dic_info['additive'] = True
            
        list_table.append(dic_info)

    # Preenche a planilha a partir da linha 4
    for baseline, data in enumerate(list_table, start=4):
        if data['additive']:
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for row in ws[f'A{baseline}:C{baseline}']:
                for cell in row:
                    cell.fill = yellow_fill

        ws.cell(row=baseline, column=1, value=data['id_teacher'])
        ws.cell(row=baseline, column=2, value=data['name_teacher'])
        ws.cell(row=baseline, column=3, value=data['monthly_hour_query'])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return current_app.response_class(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=base_pay_{year_base}_{semester_base}.xlsx'}
    )

@bp.route('/export/overtime')
@login_required
def export_excel_overtime():
    month_base = request.args.get('month_base', '')
    teacher_id = request.args.get('teacher_filter', type=int)
    
    # Regra: Se os dois estiverem vazios, não exporta
    if not month_base and not teacher_id:
        flash('Selecione pelo menos o Mês Base ou o Professor para exportar.', 'danger')
        return redirect(url_for('payments.list_overtime'))
        
    query = TeacherOvertimePay.query
    if month_base:
        query = query.filter_by(month_base=month_base)
    if teacher_id:
        query = query.filter_by(teacher_id=teacher_id)
        
    overtimes = query.all()
    
    if not overtimes:
        flash('Informações não encontradas para os filtros selecionados.', 'danger')
        return redirect(url_for('payments.list_overtime'))
        
    template_path = os.path.join(current_app.root_path, 'static', 'templates_excel', 'base_pagamento_extra.xlsx')
    
    if not os.path.exists(template_path):
        flash('Modelo do Excel não encontrado no servidor.', 'danger')
        return redirect(url_for('payments.list_overtime'))
        
    workbook = load_workbook(filename=template_path)
    ws = workbook['Extra NEB']
    
    # Preenche a data do mês base na célula D4 (apenas se o filtro de mês estiver ativo)
    if month_base:
        try:
            base_date = datetime.strptime(month_base, "%Y-%m").date()
            ws.cell(row=4, column=4, value=base_date)
        except ValueError:
            pass
            
    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'))
    font = Font(name='Calibri', size=14)
    alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    for merged_cell in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_cell))

    for baseline, data in enumerate(overtimes, start=7):
        ws.insert_rows(baseline)

        cell_data = [
            (1, data.teacher.full_name),
            (2, data.teaching_level),
            (3, data.weekly_workload),
            (4, data.hourly_value),
            (5, data.multiple_dates or ''),
            (6, data.shift),
            (7, data.budget_code),
            (8, data.justification or '')
        ]

        for col, value in cell_data:
            cell = ws.cell(row=baseline, column=col, value=value)
            cell.border = border
            cell.font = font
            cell.alignment = alignment
            
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return current_app.response_class(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=overtime_export.xlsx'}
    )