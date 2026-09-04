"""Geração da Requisição de Compra (XLSX) do módulo de Cozinha.

Soma os ingredientes das preparações selecionadas por similaridade de nome
(acentos, maiúsculas, plural e parênteses normalizados) e preenche o modelo
app/static/templates_excel/base_planilha_compras.xlsx ("REQUISIÇÃO DE COMPRA -
GASTRONOMIA"):

    Prof / Data da aula / Curso / Período
    PRODUTO | QUANTIDADE | UNIDADE | OBSERVAÇÃO   (g→KG, ml→L, un→UN)

A coluna OBSERVAÇÃO é deixada em branco, com linhas leves na tonalidade da
grade do Excel — as informações serão incluídas posteriormente. Ingredientes
de nome similar em unidades incompatíveis (g × ml × un) geram linhas
separadas por unidade. Água não é incluída na requisição.
"""
import os
import re
import unicodedata
from collections import defaultdict
from io import BytesIO

from flask import current_app
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# Modelo da requisição (mesmo padrão dos templates do módulo de Pagamentos).
TEMPLATE_PATH_PARTS = ('static', 'templates_excel', 'base_planilha_compras.xlsx')
SHEET_NAME = 'Aula dia '  # aba do modelo (com espaço no final, igual ao original)
DATA_START_ROW = 10       # primeira linha livre após o cabeçalho (linha 9)

# Linhas leves na coluna OBSERVAÇÃO, na mesma tonalidade da grade base do
# Excel — a coluna fica visualmente igual às demais e pronta para receber
# informações posteriormente.
_GRID_GRAY = 'FFD4D4D4'
_OBSERVATION_BORDER = Border(
    left=Side(style='thin', color=_GRID_GRAY),
    right=Side(style='thin', color=_GRID_GRAY),
    top=Side(style='thin', color=_GRID_GRAY),
    bottom=Side(style='thin', color=_GRID_GRAY),
)

# Produtos que não fazem sentido na requisição de compra.
EXCLUDED_INGREDIENTS = {'agua'}

# Unidades da ficha técnica → unidade base de soma e rótulo de exibição.
# (unidade normalizada) → (unidade base, fator, rótulo na planilha)
UNIT_MAP = {
    'g': ('g', 1.0, 'KG'),
    'kg': ('g', 1000.0, 'KG'),
    'mg': ('g', 0.001, 'KG'),
    'ml': ('ml', 1.0, 'L'),
    'l': ('ml', 1000.0, 'L'),
    'un': ('un', 1.0, 'UN'),
    'unid': ('un', 1.0, 'UN'),
    'unidade': ('un', 1.0, 'UN'),
    'unidades': ('un', 1.0, 'UN'),
}
DEFAULT_UNIT_LABEL = {'g': 'KG', 'ml': 'L', 'un': 'UN'}
DIVISOR = {'g': 1000.0, 'ml': 1000.0, 'un': 1.0}


def _strip_accents(text):
    return ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')


def normalize_ingredient_name(name):
    """Chave de similaridade: minúsculas, sem acentos, sem conteúdo entre
    parênteses/pontuação e no singular ('Ovos' → 'ovo')."""
    key = _strip_accents(name or '').lower()
    key = re.sub(r'\([^)]*\)|\[[^\]]*\]', ' ', key)   # parênteses/colchetes
    key = re.sub(r'[^a-z0-9]+', ' ', key).strip()
    words = key.split()
    if words and len(words[-1]) > 3 and words[-1].endswith('s'):
        words[-1] = words[-1][:-1]
    return ' '.join(words)


def _normalize_unit(unit):
    return _strip_accents((unit or '').strip().lower().rstrip('.'))


def aggregate_ingredients(recipes):
    """Soma os ingredientes das receitas por similaridade.

    Retorna lista de dicionários ordenada por nome:
        nome, quantidade (float|None), unidade (KG/L/UN/—)

    Itens sem quantidade definida ("a gosto") são agrupados no mesmo produto
    com quantidade em branco. Unidades incompatíveis do mesmo ingrediente
    (g × ml × un) geram linhas separadas por unidade.
    """
    groups = {}
    for recipe in recipes:
        for preparation in recipe.preparations:
            for ingredient in preparation.ingredients:
                # Ingredientes desativados ficam fora da requisição de compra.
                if ingredient.is_active is False:
                    continue
                key = normalize_ingredient_name(ingredient.name)
                if not key or key in EXCLUDED_INGREDIENTS:
                    continue

                group = groups.setdefault(key, {
                    'names': defaultdict(int),
                    'totals': defaultdict(float),  # unidade base → soma
                    'no_quantity': False,
                })
                group['names'][ingredient.name.strip()] += 1

                converted = UNIT_MAP.get(_normalize_unit(ingredient.unit))
                if converted and ingredient.quantity is not None:
                    base_unit, factor, _ = converted
                    group['totals'][base_unit] += ingredient.quantity * factor
                else:
                    group['no_quantity'] = True

    rows = []
    for key, group in groups.items():
        display_name = max(group['names'], key=group['names'].get)
        # Uma linha por unidade base com total (ex.: o mesmo item em g e em ml).
        unit_keys = sorted(group['totals'], key=lambda u: -group['totals'][u]) or [None]
        for row_index, base_unit in enumerate(unit_keys):
            if base_unit is not None:
                quantity = group['totals'][base_unit] / DIVISOR[base_unit]
                unit_label = DEFAULT_UNIT_LABEL[base_unit]
            elif row_index == 0:
                # Nenhum total numérico: linha única com quantidade em branco.
                quantity, unit_label = None, '—'
            else:
                break

            rows.append({
                'nome': display_name,
                'quantidade': quantity,
                'unidade': unit_label,
            })

    rows.sort(key=lambda r: _strip_accents(r['nome']).lower())
    return rows


def build_purchase_xlsx(rows, professor, class_date, course, period):
    """Preenche o modelo da requisição e retorna os bytes do arquivo XLSX.

    Levanta FileNotFoundError se o modelo base_planilha_compras.xlsx não
    estiver presente em app/static/templates_excel/.
    """
    template_path = os.path.join(current_app.root_path, *TEMPLATE_PATH_PARTS)
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f'Modelo da requisição não encontrado: {template_path}')

    workbook = load_workbook(filename=template_path)
    ws = workbook[SHEET_NAME]

    # Cabeçalho da requisição
    ws['D5'] = professor or ''
    ws['D6'] = class_date
    ws['D6'].number_format = 'dd/mm/yyyy'
    ws['A7'] = f'Curso: {course}' if course else 'Curso:'
    ws['D7'] = period or ''

    # PRODUTO | QUANTIDADE | UNIDADE — OBSERVAÇÃO (coluna 4) fica em branco,
    # com as linhas da grade para receber informações posteriormente.
    for index, row in enumerate(rows, start=DATA_START_ROW):
        ws.cell(row=index, column=1, value=row['nome'])
        if row['quantidade'] is not None:
            ws.cell(row=index, column=2, value=row['quantidade'])
        ws.cell(row=index, column=3, value=row['unidade'])
        ws.cell(row=index, column=4).border = _OBSERVATION_BORDER

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
